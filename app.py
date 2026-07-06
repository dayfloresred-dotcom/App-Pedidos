import os
from flask import Flask, render_template, request, session, redirect, url_for, jsonify, flash, Response
from datetime import date, datetime
from config import SECRET_KEY, SUCURSAL_NAMES, ADMIN_USER, now_local
from database import (init_db, crear_solicitud, get_solicitud_detalle, get_todas_solicitudes,
                       get_consolidado, get_detalle_por_sucursal, marcar_comprado, cancelar_solicitud,
                       get_db, actualizar_droguerias_pendientes,
                       get_items_detalle, marcar_item_generado, desmarcar_item_generado,
                       cancelar_producto, cancelar_producto_sucursal, cancelar_item, get_item_sucursal,
                       marcar_comprado_drogueria, marcar_inexistente,
                       registrar_envio, get_envios, get_envios_sucursal, get_envios_por_drogueria,
                       get_envios_sucursal_drogueria, omitir_drogueria, omitir_producto, restaurar_item,
                       actualizar_comprado_por_envio,
                       carrito_set, carrito_set_obs, get_carrito, carrito_clear)
from data_loader import buscar_productos, get_laboratorios, load_productos
from auth import seed_users, verify_user, login_required, admin_required
from mail_service import enviar_notificacion
from export_service import generar_suizo, generar_sud, generar_quantio

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['TEMPLATES_AUTO_RELOAD'] = True

# ── Auth ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session.get('rol') == 'admin':
        return redirect(url_for('consolidado'))
    return redirect(url_for('nueva_solicitud'))

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        result = verify_user(request.form['username'].strip(), request.form['password'])
        if result:
            session['username'], session['rol'] = result
            return redirect(url_for('index'))
        error = 'Usuario o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Sucursal + Admin ───────────────────────────────────────────────────────
@app.route('/nueva-solicitud')
@login_required
def nueva_solicitud():
    return render_template('nueva_solicitud.html',
        laboratorios=get_laboratorios(),
        sucursales=SUCURSAL_NAMES)

@app.route('/mis-pedidos')
@login_required
def mis_pedidos():
    filtro_suc = request.args.get('suc','')
    if session.get('rol') == 'admin':
        solic = get_todas_solicitudes(sucursal_filtro=filtro_suc or None)
    else:
        solic = get_todas_solicitudes(sucursal_filtro=session['username'])
    return render_template('mis_pedidos.html',
        solicitudes=solic, sucursales=SUCURSAL_NAMES, filtro_suc=filtro_suc)

@app.route('/confirmado/<int:sol_id>')
@login_required
def ver_solicitud(sol_id):
    sol, items = get_solicitud_detalle(sol_id)
    if not sol:
        flash('Solicitud no encontrada', 'danger')
        return redirect(url_for('mis_pedidos'))
    envios = get_envios_sucursal(sol['sucursal'])
    return render_template('confirmado.html', sol=sol, items=items, envios=envios,
        filtro_suc=request.args.get('suc', ''))

@app.route('/solicitud/<int:sol_id>/cancelar', methods=['POST'])
@login_required
def cancelar(sol_id):
    sol, _ = get_solicitud_detalle(sol_id)
    if not sol:
        flash('Solicitud no encontrada', 'danger')
        return redirect(url_for('mis_pedidos'))
    # Solo puede cancelar el dueño del pedido o el admin
    if session.get('rol') != 'admin' and sol['sucursal'] != session.get('username'):
        flash('No tenés permiso para cancelar este pedido', 'danger')
        return redirect(url_for('mis_pedidos'))
    if sol['estado'] != 'pendiente':
        flash('Solo se pueden cancelar pedidos pendientes', 'warning')
        return redirect(url_for('ver_solicitud', sol_id=sol_id))
    cancelar_solicitud(sol_id)
    flash(f'Pedido {sol["numero"]} cancelado', 'warning')
    return redirect(url_for('mis_pedidos'))

# ── Admin only ─────────────────────────────────────────────────────────────
@app.route('/consolidado')
@login_required
@admin_required
def consolidado():
    lab  = request.args.get('lab','')
    suc  = request.args.get('suc','')
    drog = request.args.get('drog','')
    prods = get_consolidado(
        sucursal_filtro=suc or None,
        lab_filtro=lab or None,
        drogueria_filtro=drog or None
    )
    suc_set = set()
    total_u = 0
    for p in prods:
        total_u += p['total']
        for s in (p.get('sucursales') or '').split(','):
            if s.strip(): suc_set.add(s.strip())
    return render_template('consolidado.html',
        productos=prods, sucursales=SUCURSAL_NAMES,
        n_sucursales=len(suc_set), total_unidades=total_u,
        filtro_lab=lab, filtro_suc=suc, filtro_drog=drog)

@app.route('/generar-orden')
@login_required
@admin_required
def generar_orden():
    filtro_suc = request.args.get('suc', '')
    prods    = get_consolidado(sucursal_filtro=filtro_suc or None)
    prod_map = {p['sku']: p for p in load_productos()}
    orden    = {'DROGUERIA RED': [], 'SUD': [], 'SUIZO': [], 'SIN_PRECIO': []}

    conn = get_db()
    omit_map = {}
    for r in conn.execute("SELECT sucursal, sku, drogueria FROM omitidos").fetchall():
        omit_map.setdefault((r['sucursal'], r['sku']), set()).add(r['drogueria'])
    conn.close()

    def visibles(detalle, sku, code):
        """Pills que todavía hay que gestionar en la tarjeta de esa droguería."""
        out = []
        for d in detalle:
            env = d['enviado']
            if env.get(code):
                continue          # ya enviado desde esta droguería
            if env.get('ROT'):
                continue          # cubierto por rotación
            if code in omit_map.get((d['sucursal'], sku), set()):
                continue          # quitado de esta droguería
            out.append(d)
        return out

    for p in prods:
        sku = p['sku']
        suc_list = [s.strip() for s in (p.get('sucursales') or '').split(',') if s.strip()]
        chips = ' '.join(f'<span class="chip-suc">{s}</span>' for s in suc_list[:3])
        if len(suc_list) > 3:
            chips += f' <span class="chip-suc">+{len(suc_list)-3}</span>'
        base     = prod_map.get(sku, {})
        raw_cd   = base.get('stock_cd', 0)
        stock_cd = raw_cd if isinstance(raw_cd, int) else (1 if raw_cd == 'SI' else 0)
        drog_ext = base.get('drog_ext', '')
        detalle = get_items_detalle(sku)
        if filtro_suc:
            detalle = [d for d in detalle if d['sucursal'] == filtro_suc]
        for d in detalle:
            d['enviado'] = get_envios(d['sucursal'], sku)
        drog = (p.get('drogueria') or '').upper()
        base_item = {**p, 'sucursales_str': chips, 'stock_cd': stock_cd, 'drog_ext': drog_ext}

        if drog == 'DROGUERIA RED':
            vis_cd = visibles(detalle, sku, 'CD')
            if vis_cd:
                orden['DROGUERIA RED'].append({**base_item, 'es_overflow': False,
                    'detalle_suc': vis_cd, 'drog_code': 'CD'})
            overflow = p['total'] - stock_cd
            if overflow > 0 and drog_ext in ('SUD', 'SUIZO'):
                vis_ext = visibles(detalle, sku, drog_ext)
                if vis_ext:
                    orden[drog_ext].append({**base_item, 'total': overflow, 'es_overflow': True,
                        'detalle_suc': vis_ext, 'drog_code': drog_ext})
        elif drog in ('SUD', 'SUIZO'):
            vis = visibles(detalle, sku, drog)
            if vis:
                orden[drog].append({**base_item, 'es_overflow': False,
                    'detalle_suc': vis, 'drog_code': drog})
        else:
            orden['SIN_PRECIO'].append({**base_item, 'es_overflow': False,
                'detalle_suc': detalle, 'drog_code': ''})

    conn = get_db()
    sucs_set = {r['sucursal'] for r in conn.execute("SELECT DISTINCT sucursal FROM envios WHERE cantidad>0").fetchall()}
    conn.close()
    for its in orden.values():
        for it in its:
            for d in it['detalle_suc']:
                sucs_set.add(d['sucursal'])
    sucs_orden = sorted(sucs_set)

    return render_template('generar_orden.html',
        orden={k: v for k, v in orden.items() if v},
        sucs_orden=sucs_orden,
        sucursales=SUCURSAL_NAMES,
        filtro_suc=filtro_suc,
        hoy=now_local().strftime('%d/%m/%Y'))

# ── API endpoints ──────────────────────────────────────────────────────────
@app.route('/api/productos')
@login_required
def api_productos():
    q   = request.args.get('q', '')
    lab = request.args.get('lab', '')
    suc = request.args.get('suc', '') or session.get('username', '')
    results = buscar_productos(q=q, laboratorio=lab, sucursal=suc)[:200]
    return jsonify([{
        'sku':         p['sku'],
        'ean':         p['ean'],
        'descripcion': p['descripcion'],
        'laboratorio': p['laboratorio'],
        'drogueria':   p['drogueria'],
        'ventas':      p['ventas'].get(suc, 0) if suc and suc != 'admin' else 0,
        'stock_real':  p['stock_real'].get(suc, 0) if suc and suc != 'admin' else 0,
        'stock_cd':    'SI' if (p.get('stock_cd') or 0) not in (0, '', 'NO') else 'NO',
    } for p in results])

@app.route('/api/solicitud', methods=['POST'])
@login_required
def api_crear_solicitud():
    data     = request.get_json()
    sucursal = data.get('sucursal') or session.get('username')
    items    = data.get('items', [])
    if not items:
        return jsonify({'error': 'Sin productos'}), 400
    numero, sol_id = crear_solicitud(sucursal, session['username'], items)
    enviar_notificacion(numero, sucursal, items)
    return jsonify({'numero': numero, 'sol_id': sol_id})

@app.route('/api/carrito', methods=['GET'])
@login_required
def api_carrito_get():
    suc = request.args.get('suc', '') or session.get('username', '')
    items = get_carrito(suc)
    return jsonify({'items': items, 'n': len(items)})

@app.route('/api/carrito/set', methods=['POST'])
@login_required
def api_carrito_set():
    d = request.get_json(silent=True) or {}
    suc = (d.get('sucursal') or session.get('username') or '').strip()
    if not suc:
        return jsonify({'error': 'Falta sucursal'}), 400
    carrito_set(suc, str(d.get('sku')), d.get('ean', ''), d.get('descripcion', ''),
                d.get('laboratorio', ''), d.get('drogueria', ''), d.get('cantidad', 0), d.get('observacion'))
    return jsonify({'ok': True, 'n': len(get_carrito(suc))})

@app.route('/api/carrito/obs', methods=['POST'])
@login_required
def api_carrito_obs():
    d = request.get_json(silent=True) or {}
    suc = (d.get('sucursal') or session.get('username') or '').strip()
    carrito_set_obs(suc, str(d.get('sku')), d.get('observacion') or None)
    return jsonify({'ok': True})

@app.route('/api/carrito/clear', methods=['POST'])
@login_required
def api_carrito_clear():
    d = request.get_json(silent=True) or {}
    suc = (d.get('sucursal') or session.get('username') or '').strip()
    carrito_clear(suc)
    return jsonify({'ok': True})

@app.route('/api/carrito/confirmar', methods=['POST'])
@login_required
def api_carrito_confirmar():
    d = request.get_json(silent=True) or {}
    suc = (d.get('sucursal') or session.get('username') or '').strip()
    if not suc:
        return jsonify({'error': 'Falta sucursal'}), 400
    items = get_carrito(suc)
    if not items:
        return jsonify({'error': 'El carrito está vacío'}), 400
    numero, sol_id = crear_solicitud(suc, session['username'], items)
    enviar_notificacion(numero, suc, items)
    carrito_clear(suc)
    return jsonify({'ok': True, 'sol_id': sol_id, 'numero': numero})

@app.route('/carrito')
@login_required
def carrito():
    suc = request.args.get('suc', '') or session.get('username', '')
    items = get_carrito(suc)
    total = sum(i['cantidad'] for i in items)
    return render_template('carrito.html', items=items, sucursal=suc, total=total)

@app.route('/api/detalle-sucursal')
@login_required
@admin_required
def api_detalle_sucursal():
    sku  = request.args.get('sku', '')
    rows = get_detalle_por_sucursal(sku)
    prod_map = {p['sku']: p for p in load_productos()}
    prod = prod_map.get(sku, {})
    result = []
    for r in rows:
        suc = r['sucursal']
        stock     = prod.get('stock_real', {}).get(suc, 0)
        necesidad = prod.get('necesidad', {}).get(suc, 0)
        # ventas field added in v2; fallback: necesidad + stock
        ventas_map = prod.get('ventas')
        if ventas_map is not None:
            ventas = ventas_map.get(suc, 0)
        else:
            ventas = necesidad + stock  # reconstruct: necesidad = ventas - stock
        result.append({
            'sucursal':  suc,
            'cantidad':  r['cantidad'],
            'stock':     stock,
            'ventas':    ventas,
            'necesidad': necesidad,
        })
    return jsonify(result)

@app.route('/api/orden/remove', methods=['POST'])
@login_required
@admin_required
def api_orden_remove():
    # Note: This removes pending items from display by marking them in session
    # For simplicity, we return ok and client reloads — actual removal would need
    # a separate "orden_override" table. For now this is a no-op that refreshes.
    return jsonify({'ok': True})

@app.route('/api/orden/comprado', methods=['POST'])
@login_required
@admin_required
def api_marcar_comprado():
    """Marca como comprados SOLO los productos de esa droguería (a nivel item)."""
    data  = request.get_json(silent=True) or {}
    drog  = (data.get('drogueria') or '').strip()
    fecha = data.get('fecha') or now_local().strftime('%d/%m/%Y')
    if not drog:
        return jsonify({'error': 'Falta droguería'}), 400
    n = marcar_comprado_drogueria(drog, fecha)
    return jsonify({'ok': True, 'n': n})

@app.route('/api/orden/inexistente', methods=['POST'])
@login_required
@admin_required
def api_inexistente():
    """Marca productos sin precio como inexistentes (cancelados con origen 'Producto inexistente')."""
    skus = (request.get_json(silent=True) or {}).get('skus') or []
    for sku in skus:
        sku = str(sku).strip()
        if sku:
            marcar_inexistente(sku)
    return jsonify({'ok': True, 'n': len(skus)})

@app.route('/api/orden/enviar', methods=['POST'])
@login_required
@admin_required
def api_enviar():
    """Registra cuántas unidades se envían a una sucursal desde una droguería (acumulativo)."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    drog = (data.get('drogueria') or '').strip().upper()
    try:
        cant = int(data.get('cantidad'))
    except (TypeError, ValueError):
        cant = 0
    if not (sku and suc and drog):
        return jsonify({'error': 'Faltan datos'}), 400
    registrar_envio(suc, sku, drog, cant)
    if cant > 0:
        marcar_item_generado(sku, suc, drog, now_local().strftime('%d/%m/%Y'))
    actualizar_comprado_por_envio(suc, sku)
    return jsonify({'ok': True, 'enviado': get_envios(suc, sku)})

@app.route('/api/orden/generar-item', methods=['POST'])
@login_required
@admin_required
def api_generar_item():
    """Marca como generado el pedido de un producto para una sucursal puntual."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    drog = (data.get('drogueria') or '').strip().upper()
    if not (sku and suc and drog):
        return jsonify({'error': 'Faltan datos'}), 400
    marcar_item_generado(sku, suc, drog, now_local().strftime('%d/%m/%Y'))
    return jsonify({'ok': True})

@app.route('/api/orden/desmarcar-item', methods=['POST'])
@login_required
@admin_required
def api_desmarcar_item():
    """Revierte el marcado de generado de un producto para una sucursal."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    if not (sku and suc):
        return jsonify({'error': 'Faltan datos'}), 400
    desmarcar_item_generado(sku, suc)
    return jsonify({'ok': True})

@app.route('/api/orden/cancelar-producto', methods=['POST'])
@login_required
@admin_required
def api_cancelar_producto():
    """Cancela un producto para todas las sucursales pendientes (vista compras)."""
    sku = str((request.get_json(silent=True) or {}).get('sku') or '').strip()
    if not sku:
        return jsonify({'error': 'Falta sku'}), 400
    cancelar_producto(sku)
    return jsonify({'ok': True})

@app.route('/api/orden/cancelar-item', methods=['POST'])
@login_required
@admin_required
def api_cancelar_item_suc():
    """Cancela un producto para una sucursal puntual (vista compras)."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    if not (sku and suc):
        return jsonify({'error': 'Faltan datos'}), 400
    cancelar_producto_sucursal(sku, suc)
    return jsonify({'ok': True})

@app.route('/api/item/cancelar', methods=['POST'])
@login_required
def api_cancelar_item_id():
    """Cancela un ítem puntual de un pedido (vista sucursal). Permiso: admin o dueño."""
    data = request.get_json(silent=True) or {}
    try:
        item_id = int(data.get('item_id'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Falta el ítem'}), 400
    suc = get_item_sucursal(item_id)
    if suc is None:
        return jsonify({'error': 'Ítem no encontrado'}), 404
    if session.get('rol') != 'admin' and session.get('username') != suc:
        return jsonify({'error': 'Sin permiso'}), 403
    cancelar_item(item_id)
    return jsonify({'ok': True})

@app.route('/api/orden/rotacion', methods=['POST'])
@login_required
@admin_required
def api_rotacion():
    """Marca que un producto para una sucursal se cubre por ROTACIÓN (stock interno)."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    try:
        cant = int(data.get('cantidad'))
    except (TypeError, ValueError):
        cant = 0
    if not (sku and suc):
        return jsonify({'error': 'Faltan datos'}), 400
    registrar_envio(suc, sku, 'ROT', cant)
    if cant > 0:
        marcar_item_generado(sku, suc, 'ROT', now_local().strftime('%d/%m/%Y'))
    actualizar_comprado_por_envio(suc, sku)
    return jsonify({'ok': True, 'enviado': get_envios(suc, sku)})

@app.route('/api/orden/omitir', methods=['POST'])
@login_required
@admin_required
def api_omitir():
    """Quita un producto de UNA droguería para una sucursal (se oculta de esa tarjeta)."""
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    suc  = (data.get('sucursal') or '').strip()
    drog = (data.get('drogueria') or '').strip().upper()
    if not (sku and suc and drog):
        return jsonify({'error': 'Faltan datos'}), 400
    omitir_drogueria(suc, sku, drog)
    return jsonify({'ok': True})

@app.route('/api/orden/omitir-producto', methods=['POST'])
@login_required
@admin_required
def api_omitir_producto():
    data = request.get_json(silent=True) or {}
    sku  = str(data.get('sku') or '').strip()
    drog = (data.get('drogueria') or '').strip().upper()
    if not (sku and drog):
        return jsonify({'error': 'Faltan datos'}), 400
    omitir_producto(sku, drog)
    return jsonify({'ok': True})

@app.route('/api/item/restaurar', methods=['POST'])
@login_required
def api_restaurar_item():
    """Restaura un ítem cancelado (vuelve a Generar orden). Permiso: admin o dueño."""
    data = request.get_json(silent=True) or {}
    try:
        item_id = int(data.get('item_id'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Falta el ítem'}), 400
    suc = get_item_sucursal(item_id)
    if suc is None:
        return jsonify({'error': 'Ítem no encontrado'}), 404
    if session.get('rol') != 'admin' and session.get('username') != suc:
        return jsonify({'error': 'Sin permiso'}), 403
    restaurar_item(item_id)
    return jsonify({'ok': True})

# ── Export routes ──────────────────────────────────────────────────────────
@app.route('/exportar/<drogueria>', methods=['GET', 'POST'])
@login_required
@admin_required
def exportar(drogueria):
    """Exporta el pedido de UNA sucursal para esa droguería, con las cantidades manuales cargadas."""
    drog = drogueria.upper()
    data = request.get_json(silent=True) or {}
    sucursal = (data.get('sucursal') or '').strip()
    if not sucursal:
        return jsonify({'error': 'Elegí una sucursal'}), 400

    prod_map = {p['sku']: p for p in load_productos()}
    items = []
    for r in get_envios_sucursal_drogueria(sucursal, drog):
        base = prod_map.get(r['sku'], {})
        items.append({
            'ean':         base.get('ean', ''),
            'troquel':     base.get('troquel', '0000000'),
            'descripcion': base.get('descripcion', ''),
            'cantidad':    r['cantidad'],
        })
    if not items:
        return jsonify({'error': f'{sucursal} no tiene envíos cargados para {drog}.'}), 400

    suc_slug = ''.join(c if c.isalnum() else '_' for c in sucursal.lower())
    if drog == 'SUIZO':
        content  = generar_suizo(items)
        filename = f'suizo_{suc_slug}_{now_local().strftime("%d%m%y")}.arg'
    else:
        content  = generar_sud(items)
        filename = f'sud_{suc_slug}_{now_local().strftime("%d%m%y")}.dds'

    return Response(
        content,
        mimetype='application/octet-stream',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.route('/exportar-quantio', methods=['POST'])
@login_required
@admin_required
def exportar_quantio():
    """Pedido Quantio (CD) por sucursal con cantidades editadas."""
    data     = request.get_json(silent=True) or {}
    sucursal = (data.get('sucursal') or '').strip()
    if not sucursal:
        return jsonify({'error': 'Falta sucursal'}), 400

    prod_map = {p['sku']: p for p in load_productos()}
    items = []
    for r in get_envios_sucursal_drogueria(sucursal, 'CD'):
        base = prod_map.get(r['sku'], {})
        items.append({
            'ean':     base.get('ean', ''),
            'troquel': base.get('troquel_pres') or base.get('troquel') or '',
            'cantidad': r['cantidad'],
        })

    if not items:
        return jsonify({'error': f'{sucursal} no tiene envíos cargados para CD.'}), 400

    content   = generar_quantio(items)
    suc_slug  = ''.join(c if c.isalnum() else '_' for c in sucursal.lower())
    filename  = f'quantio_{suc_slug}_{now_local().strftime("%d%m%y")}.txt'
    return Response(
        content.encode('latin-1', errors='replace'),
        mimetype='text/plain',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.route('/reporte/rotacion.xlsx')
@login_required
@admin_required
def reporte_rotacion():
    """Excel con movimientos de rotación sugeridos: de sucursales con stock alto y 0 ventas
    hacia sucursales con stock 0 que sí venden."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    UMBRAL = 2  # 'stock alto' del donante
    prods = load_productos()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Rotacion'
    ws.append(['EAN', 'Producto', 'Laboratorio', 'Desde (sucursal)', 'Stock donante',
               'Hacia (sucursal)', 'Ventas receptor', 'Cantidad a mover'])
    for p in prods:
        stock  = p.get('stock_real', {}) or {}
        ventas = p.get('ventas', {}) or {}
        sucs = set(stock) | set(ventas)
        donors = [(su, stock.get(su, 0)) for su in sucs if stock.get(su, 0) >= UMBRAL and ventas.get(su, 0) == 0]
        recept = [(su, ventas.get(su, 0)) for su in sucs if stock.get(su, 0) == 0 and ventas.get(su, 0) > 0]
        if not donors or not recept:
            continue
        donors = sorted(donors, key=lambda x: -x[1])
        recept = sorted(recept, key=lambda x: -x[1])
        avail = [[su, q] for su, q in donors]          # stock disponible mutable
        orig  = {su: q for su, q in donors}            # stock original (para mostrar)
        for rsuc, rneed in recept:
            need = rneed
            for row in avail:
                if need <= 0:
                    break
                if row[1] <= 0:
                    continue
                move = min(row[1], need)
                ws.append([p['ean'], (p['descripcion'] or '')[:45], p['laboratorio'],
                           row[0], orig[row[0]], rsuc, rneed, move])
                row[1] -= move
                need   -= move
    # estilo cabecera
    fill = PatternFill('solid', start_color='1F4E78')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = fill; c.alignment = Alignment(horizontal='center')
    for col, w in {'A':15,'B':45,'C':22,'D':16,'E':13,'F':16,'G':14,'H':15}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f'rotacion_{now_local().strftime("%d%m%y")}.xlsx'
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})

@app.route('/actualizar-datos', methods=['GET', 'POST'])
@login_required
@admin_required
def actualizar_datos():
    from config import (PRESUPUESTO_CSV, STOCK_CD_CSV, LISTADO_STOCK_CSV,
                        PRECIOS_SUD_TXT, PRECIOS_SUIZO_PERFU, PRECIOS_SUIZO_INS,
                        PRECIOS_DDS_XLSX, PRECIOS_SUIZO_CMP_XLSX)
    import data_loader

    ARCHIVOS = {
        'presupuesto':         {'label': 'Presupuesto (ventas, stock y stock CD)',  'path': PRESUPUESTO_CSV,       'accept': '.csv,.txt', 'diario': True},
        'precios_dds':         {'label': 'Lista DDS / Sud (comparador, precio final)', 'path': PRECIOS_DDS_XLSX,    'accept': '.xlsx',     'diario': False},
        'precios_suizo_cmp':   {'label': 'Lista Suizo (comparador, precio final)',   'path': PRECIOS_SUIZO_CMP_XLSX,'accept': '.xlsx',     'diario': False},
        'precios_sud':         {'label': 'Precios SUD (formato viejo .txt)',         'path': PRECIOS_SUD_TXT,       'accept': '.txt',      'diario': False},
        'precios_suizo_perfu': {'label': 'Precios Suizo Perfumería (viejo)',         'path': PRECIOS_SUIZO_PERFU,   'accept': '.xls,.xlsx','diario': False},
        'precios_suizo_ins':   {'label': 'Precios Suizo Insumos (viejo)',            'path': PRECIOS_SUIZO_INS,     'accept': '.xls,.xlsx','diario': False},
    }

    if request.method == 'POST':
        updated = []
        for field, cfg in ARCHIVOS.items():
            f = request.files.get(field)
            if f and f.filename:
                os.makedirs(os.path.dirname(cfg['path']), exist_ok=True)
                f.save(cfg['path'])
                updated.append(cfg['label'])
        if updated:
            # Forzar recarga completa: borrar cache en disco y en memoria
            if os.path.exists(data_loader.CACHE_FILE):
                os.remove(data_loader.CACHE_FILE)
            data_loader._productos = None
            nuevos = load_productos()
            prod_map = {p['sku']: p for p in nuevos}
            n_items = actualizar_droguerias_pendientes(prod_map)
            flash(f'✓ Actualizados: {", ".join(updated)}. {len(nuevos)} productos recargados. {n_items} pedidos pendientes actualizados.', 'success')
        else:
            flash('No seleccionaste ningún archivo.', 'warning')
        return redirect(url_for('actualizar_datos'))

    archivos_info = {}
    for field, cfg in ARCHIVOS.items():
        path = cfg['path']
        if os.path.exists(path):
            mtime = os.path.getmtime(path)
            info = {'existe': True, 'fecha': datetime.fromtimestamp(mtime).strftime('%d/%m/%Y %H:%M')}
        else:
            info = {'existe': False, 'fecha': '—'}
        archivos_info[field] = {**cfg, **info, 'nombre': os.path.basename(path)}
    return render_template('actualizar_datos.html', archivos=archivos_info)

@app.errorhandler(403)
def forbidden(e):
    return render_template('login.html', error='Acceso denegado'), 403

init_db()
seed_users()

if __name__ == '__main__':
    print("Iniciando App Pedidos Farmacias Red...")
    load_productos()
    print("Listo! Abrí tu navegador en: http://localhost:8080")
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=8080)

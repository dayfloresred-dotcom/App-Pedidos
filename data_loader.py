import csv, re, os, pickle
from config import (PRESUPUESTO_CSV, LISTADO_STOCK_CSV, STOCK_CD_CSV,
                    PRECIOS_SUD_TXT, PRECIOS_SUIZO_PERFU, PRECIOS_SUIZO_INS,
                    PRECIOS_DDS_XLSX, PRECIOS_SUIZO_CMP_XLSX,
                    SUCURSALES, BASE_DIR)

RUBROS   = {'Perfumería', 'Accesorios'}
EXCLUIR  = {'17', '33'}
# v2: agrega troquel_pres (Troquel del presupuesto) — el cambio de nombre fuerza regenerar el cache
CACHE_FILE = os.path.join(BASE_DIR, 'productos_cache_v3.pkl')
SOURCE_FILES = [PRESUPUESTO_CSV, LISTADO_STOCK_CSV, STOCK_CD_CSV,
                PRECIOS_SUD_TXT, PRECIOS_SUIZO_PERFU, PRECIOS_SUIZO_INS,
                PRECIOS_DDS_XLSX, PRECIOS_SUIZO_CMP_XLSX]

_productos = None   # list of dicts, loaded once at startup

def _cache_valid():
    if not os.path.exists(CACHE_FILE):
        return False
    cache_mtime = os.path.getmtime(CACHE_FILE)
    return all(os.path.getmtime(f) <= cache_mtime for f in SOURCE_FILES if os.path.exists(f))

def _extract_num(s):
    return s.replace('Suc. ','').replace('SUC','').strip()

def _load_eans():
    if not os.path.exists(LISTADO_STOCK_CSV):
        return {}
    eans = {}
    with open(LISTADO_STOCK_CSV, encoding='latin-1') as f:
        for i, line in enumerate(f):
            if i < 12: continue
            row = line.rstrip().split(';')
            if len(row) > 2 and row[0].strip():
                eans[row[0].strip()] = row[2].strip()
    return eans

def _load_cd_stock():
    """Returns dict: sku -> cantidad en CD (solo productos con cantidad > 0)"""
    if not os.path.exists(STOCK_CD_CSV):
        return {}
    cd = {}
    with open(STOCK_CD_CSV, encoding='latin-1') as f:
        for row in csv.DictReader(f, delimiter=';'):
            try:
                cant = float(row['Cantidad'].replace(',','.'))
                if cant > 0:
                    cd[row['idProducto'].strip()] = int(cant)
            except: pass
    return cd

def _load_sud_prices():
    if not os.path.exists(PRECIOS_SUD_TXT):
        return {}
    prices = {}
    for enc in ['utf-8-sig', 'latin-1']:
        try:
            with open(PRECIOS_SUD_TXT, encoding=enc) as f:
                for line in f:
                    if not line.startswith('D'): continue
                    eans = re.findall(r'HE(\d{13})', line)
                    try: price = int(line[163:175]) / 100
                    except: continue
                    if price <= 0: continue
                    for ean in eans:
                        prices[ean] = price
            break
        except (UnicodeDecodeError, ValueError):
            prices = {}
            continue
    return prices

def _load_suizo_prices():
    prices = {}
    for fname in [PRECIOS_SUIZO_PERFU, PRECIOS_SUIZO_INS]:
        if not os.path.exists(fname):
            continue
        with open(fname, encoding='latin-1') as f:
            for i, line in enumerate(f):
                if i == 0: continue
                cols = line.rstrip().split('\t')
                if len(cols) < 5: continue
                try:
                    p = float((cols[4] or cols[3]).strip('"').replace(',','.').strip())
                except: continue
                if p < 1: continue
                for c in [1, 13, 14, 15, 16, 17]:
                    if c < len(cols):
                        ean = cols[c].strip('"').strip()
                        if ean and len(ean) >= 7:
                            prices[ean] = p
    return prices

def _load_dds_prices():
    """Reporte DDS (Sud) del comparador -> {EAN: precio final con IVA}.
    Precio base = 'P. c/dtos + 12% (NM)'. Medicamentos (ME) sin IVA; resto +21%."""
    if not os.path.exists(PRECIOS_DDS_XLSX):
        return {}
    prices = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PRECIOS_DDS_XLSX, read_only=True, data_only=True)
        ws = wb.active
        idx = None
        for row in ws.iter_rows(values_only=True):
            cells = [('' if c is None else str(c).strip()) for c in row]
            if idx is None:
                if 'EAN' in cells:
                    idx = {}
                    for i, c in enumerate(cells):
                        cl = c.lower()
                        if c == 'EAN': idx['ean'] = i
                        elif 'c/dtos + 12%' in cl: idx['neto'] = i
                        elif 'categor' in cl: idx['cat'] = i
                continue
            if 'ean' not in idx or 'neto' not in idx:
                break
            try:
                ean = str(int(float(cells[idx['ean']])))
            except (ValueError, TypeError):
                continue
            try:
                neto = float(row[idx['neto']])
            except (ValueError, TypeError):
                continue
            cat = cells[idx['cat']].upper() if 'cat' in idx else ''
            final = neto if cat == 'ME' else neto * 1.21
            if final > 0:
                prices[ean] = round(final, 2)
        wb.close()
    except Exception as e:
        print('[DATA] Error leyendo reporte DDS comparador:', e)
        return {}
    return prices

def _load_suizo_cmp_prices():
    """Reporte Suizo del comparador -> {EAN: precio final con IVA}.
    Base = 'Su Precio'; x0.88 si 'Aplica 12%'=Si; x1.21 si 'Aplica IVA'=Si."""
    if not os.path.exists(PRECIOS_SUIZO_CMP_XLSX):
        return {}
    prices = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PRECIOS_SUIZO_CMP_XLSX, read_only=True, data_only=True)
        ws = wb.active
        idx = None
        for row in ws.iter_rows(values_only=True):
            cells = [('' if c is None else str(c).strip()) for c in row]
            if idx is None:
                if 'EAN' in cells:
                    idx = {}
                    for i, c in enumerate(cells):
                        cl = c.lower()
                        if c == 'EAN': idx['ean'] = i
                        elif cl == 'su precio': idx['precio'] = i
                        elif 'aplica 12' in cl: idx['a12'] = i
                        elif 'aplica iva' in cl: idx['aiva'] = i
                continue
            if 'ean' not in idx or 'precio' not in idx:
                break
            try:
                ean = str(int(float(cells[idx['ean']])))
            except (ValueError, TypeError):
                continue
            try:
                base = float(row[idx['precio']])
            except (ValueError, TypeError):
                continue
            if 'a12' in idx and cells[idx['a12']].lower() == 'si':
                base *= 0.88
            if 'aiva' in idx and cells[idx['aiva']].lower() == 'si':
                base *= 1.21
            if base > 0:
                prices[ean] = round(base, 2)
        wb.close()
    except Exception as e:
        print('[DATA] Error leyendo reporte Suizo comparador:', e)
        return {}
    return prices

def _load_troqueles_sud():
    """Map EAN -> troquel (7 chars) from SUD price file for .dds export."""
    if not os.path.exists(PRECIOS_SUD_TXT):
        return {}
    troqueles = {}
    for enc in ['utf-8-sig', 'latin-1']:
        try:
            with open(PRECIOS_SUD_TXT, encoding=enc) as f:
                for line in f:
                    if not line.startswith('D'): continue
                    troquel = str(int(line[1:19]))[-7:].zfill(7)
                    eans = re.findall(r'HE(\d{13})', line)
                    for ean in eans:
                        troqueles[ean] = troquel
            break
        except (UnicodeDecodeError, ValueError):
            troqueles = {}
            continue
    return troqueles

def load_productos():
    global _productos
    if _productos is not None:
        return _productos

    # Try loading from cache first
    if _cache_valid():
        print('[DATA] Cargando desde cache...')
        with open(CACHE_FILE, 'rb') as f:
            _productos = pickle.load(f)
        print(f'[DATA] {len(_productos)} productos cargados desde cache (rápido)')
        return _productos

    # Si no existe el archivo principal, retornar lista vacía
    if not os.path.exists(PRESUPUESTO_CSV):
        print('[DATA] Archivo de presupuesto no encontrado. Subí los archivos desde "Actualizar datos".')
        _productos = []
        return _productos

    print('[DATA] Cargando productos desde archivos CSV (primera vez, puede tardar)...')
    if _productos is not None:
        return _productos

    eans      = _load_eans()
    cd_stock  = _load_cd_stock()
    sud_cmp   = _load_dds_prices()
    suizo_cmp = _load_suizo_cmp_prices()
    sud_p     = sud_cmp if sud_cmp else _load_sud_prices()
    suizo_p   = suizo_cmp if suizo_cmp else _load_suizo_prices()
    if sud_cmp:   print(f'[DATA] Precios SUD desde comparador: {len(sud_cmp)}')
    if suizo_cmp: print(f'[DATA] Precios SUIZO desde comparador: {len(suizo_cmp)}')
    troqueles = _load_troqueles_sud()

    headers     = []
    header_row  = -1
    suc_vend    = {}
    suc_stock   = {}
    cd_col_idx  = None   # índice de columna "Cajas Stock CD" en el presupuesto
    troquel_idx = None   # índice de columna "Troquel" en el presupuesto

    productos = []
    with open(PRESUPUESTO_CSV, encoding='latin-1') as f:
        reader = csv.reader(f, delimiter=';')
        for i, row in enumerate(reader):
            # Find header row dynamically (contains 'IDProducto')
            if not headers and row and row[0].strip() == 'IDProducto':
                headers    = row
                header_row = i
                for idx, h in enumerate(headers):
                    h_strip = h.strip()
                    if h_strip == 'Cajas Stock CD':
                        cd_col_idx = idx
                    elif h_strip == 'Troquel':
                        troquel_idx = idx
                    elif h.startswith('Cajas Vend'):
                        num = _extract_num(h.replace('Cajas Vend. ','').replace('Cajas Vend ',''))
                        if num not in EXCLUIR and num in SUCURSALES:
                            suc_vend[SUCURSALES[num]] = idx
                    elif h.startswith('Cajas Stock'):
                        num = _extract_num(h.replace('Cajas Stock ',''))
                        if num not in EXCLUIR and num in SUCURSALES:
                            suc_stock[SUCURSALES[num]] = idx
            elif header_row >= 0 and i > header_row and row and len(row) > 5 and row[3].strip():
                sku = row[0].strip()
                # Use EAN from new file's Codebar column first (96% coverage, no scientific notation)
                # Fall back to Listado de Stock lookup, then skip if scientific notation
                ean_csv = row[2].strip()
                if ean_csv and ean_csv != '0' and 'E+' not in ean_csv and 'e+' not in ean_csv:
                    ean = ean_csv
                else:
                    ean = eans.get(sku, '')

                # Troquel del presupuesto (para export Quantio / distribución CD)
                if troquel_idx is not None and troquel_idx < len(row):
                    troquel_pres = row[troquel_idx].strip()
                elif len(row) > 1:
                    troquel_pres = row[1].strip()
                else:
                    troquel_pres = ''

                # Stock CD: preferir columna del presupuesto, fallback al archivo Stock CD.csv
                if cd_col_idx is not None and cd_col_idx < len(row):
                    try:
                        cant_cd = int(float(row[cd_col_idx].replace(',', '.'))) if row[cd_col_idx].strip() else 0
                        cant_cd = max(0, cant_cd)
                    except:
                        cant_cd = cd_stock.get(sku, 0)
                else:
                    cant_cd = cd_stock.get(sku, 0)
                tiene_cd = cant_cd > 0

                # Always compute external droguería (fallback if CD stock insufficient)
                p_sud   = sud_p.get(ean)
                p_suizo = suizo_p.get(ean)
                if p_sud and p_suizo:
                    drog_ext = 'SUD' if p_sud <= p_suizo else 'SUIZO'
                    precio_ext = min(p_sud, p_suizo)
                elif p_sud:
                    drog_ext, precio_ext = 'SUD', p_sud
                elif p_suizo:
                    drog_ext, precio_ext = 'SUIZO', p_suizo
                else:
                    drog_ext, precio_ext = '', None

                if tiene_cd:
                    drogueria    = 'DROGUERIA RED'
                    mejor_precio = None
                else:
                    drogueria    = drog_ext
                    mejor_precio = precio_ext

                suc_data       = {}   # necesidad (ventas - stock)
                suc_stock_real = {}   # stock actual
                suc_ventas     = {}   # ventas del mes
                for sname, v_idx in suc_vend.items():
                    s_idx = suc_stock.get(sname)
                    try: ventas = int(float(row[v_idx].replace(',','.'))) if row[v_idx].strip() else 0
                    except: ventas = 0
                    try: stock = int(float(row[s_idx].replace(',','.'))) if s_idx and row[s_idx].strip() else 0
                    except: stock = 0
                    suc_data[sname]       = max(0, ventas - stock)
                    suc_stock_real[sname] = stock
                    suc_ventas[sname]     = ventas

                productos.append({
                    'sku':          sku,
                    'ean':          ean,
                    'descripcion':  row[4],
                    'laboratorio':  row[5],
                    'rubro':        row[3],
                    'stock_cd':     cant_cd,
                    'drogueria':    drogueria,
                    'mejor_precio': mejor_precio,
                    'drog_ext':     drog_ext,     # best external droguería (SUD/SUIZO)
                    'troquel':      troqueles.get(ean, '0000000'),
                    'troquel_pres': troquel_pres, # Troquel del presupuesto (export Quantio)
                    'necesidad':    suc_data,
                    'stock_real':   suc_stock_real,
                    'ventas':       suc_ventas,
                })

    _productos = productos

    # Save cache for next startup
    try:
        with open(CACHE_FILE, 'wb') as f:
            pickle.dump(_productos, f)
        print(f'[DATA] Cache guardado ({len(_productos)} productos). Próximo inicio será rápido.')
    except Exception as e:
        print(f'[DATA] No se pudo guardar cache: {e}')

    return _productos

def buscar_productos(q='', laboratorio='', sucursal=None):
    """Return filtered product list. sucursal arg used for context only."""
    prods = load_productos()
    result = []
    q   = q.lower().strip()
    lab = laboratorio.lower().strip()
    for p in prods:
        if lab and lab not in p['laboratorio'].lower():
            continue
        if q and q not in p['descripcion'].lower() and q not in p['ean']:
            continue
        result.append({**p, 'necesidad_suc': p['necesidad'].get(sucursal, 0) if sucursal else 0})
    return result

def get_laboratorios():
    prods = load_productos()
    return sorted(set(p['laboratorio'] for p in prods if p['laboratorio']))
